import json
import math
import time
import urllib.parse
import urllib.request
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data" / "english-hindi"
OUTPUT_DIR = ROOT / "outputs" / "assignment2_python"
EXCEL_FILE = OUTPUT_DIR / "assessment2_model_translations.xlsx"
METRICS_FILE = OUTPUT_DIR / "translation_metrics.txt"
MODEL_NAME = "Google Translate hosted translation model"
TRANSLATE_URL = "https://translate.googleapis.com/translate_a/single"
SENTENCE_LIMIT = 100


def word_count(sentence: str) -> int:
    sentence = sentence.strip()
    return len(sentence.split()) if sentence else 0


def read_lines(path: Path) -> list[str]:
    return path.read_text(encoding="utf-8").lstrip("\ufeff").splitlines()


def get_first_100_cleaned_rows() -> list[dict[str, str]]:
    english_lines = read_lines(DATA_DIR / "eng.txt")
    hindi_lines = read_lines(DATA_DIR / "hin.txt")

    rows = []
    for english, reference_hindi in zip(english_lines, hindi_lines):
        english = english.strip()
        reference_hindi = reference_hindi.strip()
        english_wc = word_count(english)
        hindi_wc = word_count(reference_hindi)
        difference = english_wc - hindi_wc

        if 5 <= english_wc <= 50 and 5 <= hindi_wc <= 50 and -10 <= difference <= 10:
            rows.append({"english": english, "reference_hindi": reference_hindi})

        if len(rows) == SENTENCE_LIMIT:
            break

    return rows


def translate_to_hindi(sentence: str) -> str:
    query = urllib.parse.urlencode(
        {
            "client": "gtx",
            "sl": "en",
            "tl": "hi",
            "dt": "t",
            "q": sentence,
        }
    )
    request = urllib.request.Request(f"{TRANSLATE_URL}?{query}", method="GET")
    with urllib.request.urlopen(request, timeout=90) as response:
        result = json.loads(response.read().decode("utf-8"))
    return "".join(part[0] for part in result[0] if part and part[0]).strip()


def ngrams(tokens: list[str], n: int) -> Counter:
    return Counter(tuple(tokens[i : i + n]) for i in range(max(0, len(tokens) - n + 1)))


def corpus_bleu(predictions: list[str], references: list[str], max_order: int = 4) -> float:
    matches = [0] * max_order
    totals = [0] * max_order
    pred_len = 0
    ref_len = 0

    for pred, ref in zip(predictions, references):
        pred_tokens = pred.split()
        ref_tokens = ref.split()
        pred_len += len(pred_tokens)
        ref_len += len(ref_tokens)

        for n in range(1, max_order + 1):
            pred_ngrams = ngrams(pred_tokens, n)
            ref_ngrams = ngrams(ref_tokens, n)
            matches[n - 1] += sum((pred_ngrams & ref_ngrams).values())
            totals[n - 1] += sum(pred_ngrams.values())

    precisions = [(matches[i] + 1) / (totals[i] + 1) for i in range(max_order)]
    geo_mean = math.exp(sum(math.log(p) for p in precisions) / max_order)
    brevity_penalty = 1.0 if pred_len > ref_len else math.exp(1 - ref_len / max(pred_len, 1))
    return 100 * brevity_penalty * geo_mean


def char_ngrams(text: str, n: int) -> Counter:
    compact = " ".join(text.split())
    return Counter(compact[i : i + n] for i in range(max(0, len(compact) - n + 1)))


def corpus_chrf(predictions: list[str], references: list[str], max_order: int = 6, beta: int = 2) -> float:
    beta_squared = beta * beta
    scores = []

    for n in range(1, max_order + 1):
        overlap = 0
        pred_total = 0
        ref_total = 0

        for pred, ref in zip(predictions, references):
            pred_counts = char_ngrams(pred, n)
            ref_counts = char_ngrams(ref, n)
            overlap += sum((pred_counts & ref_counts).values())
            pred_total += sum(pred_counts.values())
            ref_total += sum(ref_counts.values())

        precision = overlap / pred_total if pred_total else 0
        recall = overlap / ref_total if ref_total else 0

        if precision + recall == 0:
            scores.append(0)
        else:
            scores.append((1 + beta_squared) * precision * recall / (beta_squared * precision + recall))

    return 100 * sum(scores) / len(scores)


def edit_distance(source: list[str], target: list[str]) -> int:
    previous = list(range(len(target) + 1))
    for i, source_token in enumerate(source, start=1):
        current = [i]
        for j, target_token in enumerate(target, start=1):
            cost = 0 if source_token == target_token else 1
            current.append(min(previous[j] + 1, current[j - 1] + 1, previous[j - 1] + cost))
        previous = current
    return previous[-1]


def corpus_ter(predictions: list[str], references: list[str]) -> float:
    edits = 0
    reference_words = 0
    for pred, ref in zip(predictions, references):
        pred_tokens = pred.split()
        ref_tokens = ref.split()
        edits += edit_distance(pred_tokens, ref_tokens)
        reference_words += len(ref_tokens)
    return 100 * edits / max(reference_words, 1)


def save_excel(rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Model Translations"
    ws.append(["Original English sentence", "Model-generated Hindi translation"])

    for row in rows:
        ws.append([row["english"], row["model_hindi"]])

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 85
    ws.column_dimensions["B"].width = 85
    ws.freeze_panes = "A2"

    table = Table(displayName="Assessment2Translations", ref=f"A1:B{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    wb.save(EXCEL_FILE)


def main() -> None:
    rows = get_first_100_cleaned_rows()

    for index, row in enumerate(rows, start=1):
        row["model_hindi"] = translate_to_hindi(row["english"])
        print(f"Translated {index}/{len(rows)}")
        time.sleep(0.2)

    predictions = [row["model_hindi"] for row in rows]
    references = [row["reference_hindi"] for row in rows]

    bleu = corpus_bleu(predictions, references)
    chrf = corpus_chrf(predictions, references)
    ter = corpus_ter(predictions, references)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    save_excel(rows)

    METRICS_FILE.write_text(
        "\n".join(
            [
                "Assessment No. 2 - Translation Metrics",
                f"Model: {MODEL_NAME}",
                f"Sentence count: {len(rows)}",
                f"BLEU: {bleu:.4f}",
                f"CHRF: {chrf:.4f}",
                f"TER: {ter:.4f}",
                "",
                "Reference translations are the Hindi sentences from the cleaned Assignment No. 1 dataset.",
                "TER is computed as word-level edit distance divided by reference word count.",
            ]
        ),
        encoding="utf-8",
    )

    print(f"Saved Assignment 2 Excel file: {EXCEL_FILE}")
    print(f"Saved metrics file: {METRICS_FILE}")


if __name__ == "__main__":
    main()
