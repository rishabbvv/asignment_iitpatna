from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data" / "english-hindi"
OUTPUT_DIR = ROOT / "outputs" / "assignment1_python"
OUTPUT_FILE = OUTPUT_DIR / "english_hindi_cleaned_dataset.xlsx"


def word_count(sentence: str) -> int:
    sentence = sentence.strip()
    return len(sentence.split()) if sentence else 0


def read_lines(path: Path) -> list[str]:
    return path.read_text(encoding="utf-8").lstrip("\ufeff").splitlines()


def main() -> None:
    english_lines = read_lines(DATA_DIR / "eng.txt")
    hindi_lines = read_lines(DATA_DIR / "hin.txt")

    if len(english_lines) != len(hindi_lines):
        raise ValueError(f"Line mismatch: {len(english_lines)} English vs {len(hindi_lines)} Hindi")

    if len(english_lines) < 10000:
        raise ValueError("Dataset must contain at least 10,000 rows.")

    cleaned_rows = []
    for english, hindi in zip(english_lines, hindi_lines):
        english = english.strip()
        hindi = hindi.strip()
        english_wc = word_count(english)
        hindi_wc = word_count(hindi)
        difference = english_wc - hindi_wc

        if 5 <= english_wc <= 50 and 5 <= hindi_wc <= 50 and -10 <= difference <= 10:
            cleaned_rows.append([english, hindi, english_wc, hindi_wc, difference])

    wb = Workbook()
    ws = wb.active
    ws.title = "Cleaned Dataset"

    headers = [
        "English Sentences",
        "Hindi Sentences",
        "Word Count (English)",
        "Word Count (Hindi)",
        "Difference between Word Count (English) and Word Count (Hindi)",
    ]
    ws.append(headers)

    for row in cleaned_rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for row in ws.iter_rows(min_row=2):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        for cell in row[2:]:
            cell.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 45
    ws.freeze_panes = "A2"

    table_ref = f"A1:E{ws.max_row}"
    table = Table(displayName="CleanedEnglishHindi", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    summary = wb.create_sheet("Summary")
    summary_rows = [
        ["Original paired rows", len(english_lines)],
        ["Rows retained after filters", len(cleaned_rows)],
        ["Sentence word count rule", "Both English and Hindi word counts between 5 and 50 inclusive"],
        ["Difference rule", "English word count minus Hindi word count between -10 and +10 inclusive"],
        ["Source", "https://huggingface.co/datasets/ainlpml/english-hindi"],
    ]
    summary.append(["English-Hindi Dataset Cleaning Summary", ""])
    for row in summary_rows:
        summary.append(row)
    summary["A1"].fill = header_fill
    summary["A1"].font = header_font
    summary.column_dimensions["A"].width = 35
    summary.column_dimensions["B"].width = 90

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"Saved Assignment 1 Excel file: {OUTPUT_FILE}")
    print(f"Original rows: {len(english_lines)}")
    print(f"Rows retained: {len(cleaned_rows)}")


if __name__ == "__main__":
    main()
